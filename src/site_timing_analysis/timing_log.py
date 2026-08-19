# Project: Site Timing Analysis
# File: src/site_timing_analysis/timing_log.py
# Primary author: Nicholas J. Sisco, Ph.D.
# Organization: Profound Medical, LLC
# Created: 2026-03-11
# Purpose: Finds and parses optional timing-log CSV/XLSX files for timing enrichment.
#
# Provenance: Original implementation or material contribution by
# Nicholas J. Sisco, Ph.D. for Profound Medical, LLC.
#
# Rights status: Proprietary / internal use unless otherwise specified
# by Profound Medical, LLC.
from __future__ import annotations

import csv
import warnings as runtime_warnings
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException

from .errors import TimingLogParseError
from .models import TimingLogEntry


_EXPLICIT_COLUMNS = ("Events", "TimeSTART", "TimeEND")
_POS_EVENT_IDX = 2
_POS_START_IDX = 3
_POS_END_IDX = 4
_SUPPORTED_SUFFIXES = (".csv", ".xlsx")
_XLSX_HEADER_NAMES = ("EVENT", "START", "END")
_XLSX_SCAN_ROWS = 100
_XLSX_SCAN_COLUMNS = 64


def resolve_timing_log(
    case_id: str,
    resolved_site_root: Path,
    timing_log_dir_override: Path | None = None,
) -> tuple[Path | None, list[str]]:
    """
    Resolve one optional timing log by exact case-ID filename.

    Input:
        A canonical case ID, resolved site root, and optional shared timing-log
        directory override.
    Output:
        The sole exact ``<case_id>.csv`` or ``<case_id>.xlsx`` match plus
        recoverable resolution warnings.
    Assumptions:
        Filename normalization is intentionally not fuzzy. Multiple exact
        supported files are ambiguous and fail explicitly; a missing file is
        reported only when a timing-log directory was explicitly selected or
        the conventional site-local directory exists.
    """
    base_dir = timing_log_dir_override if timing_log_dir_override is not None else resolved_site_root / "TimingLogs"
    should_report_missing = timing_log_dir_override is not None or base_dir.is_dir()
    if not base_dir.is_dir():
        warning = (
            [f"{case_id}:timing_log_directory_missing:{base_dir.resolve()}"]
            if should_report_missing
            else []
        )
        return None, warning

    candidates = [
        candidate.resolve()
        for suffix in _SUPPORTED_SUFFIXES
        if (candidate := base_dir / f"{case_id}{suffix}").is_file()
    ]
    if len(candidates) > 1:
        raise TimingLogParseError(
            case_id,
            base_dir,
            "Ambiguous timing-log match. Expected exactly one supported file; "
            f"found: {', '.join(str(path) for path in candidates)}",
        )
    if candidates:
        return candidates[0], []
    return None, [
        f"{case_id}:timing_log_missing:{base_dir.resolve()}:"
        f"expected={case_id}.csv|{case_id}.xlsx"
    ]


def find_timing_log(
    case_id: str,
    resolved_site_root: Path,
    timing_log_dir_override: Path | None = None,
) -> Path | None:
    """Return the exact optional timing-log path without resolution warnings."""
    path, _ = resolve_timing_log(
        case_id,
        resolved_site_root,
        timing_log_dir_override,
    )
    return path


def _parse_optional_datetime(
    raw_value: str | None,
    *,
    case_id: str,
    source_path: Path,
    row_number: int,
    field_name: str,
) -> tuple[datetime | None, str | None]:
    if raw_value is None:
        return None, None

    text = raw_value.strip()
    if text == "":
        return None, None

    normalized = text.replace("T", " ")
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    if "." in normalized:
        left, right = normalized.split(".", 1)
        frac = right
        tz_suffix = ""
        if "+" in right:
            frac, tz_suffix = right.split("+", 1)
            tz_suffix = "+" + tz_suffix
        elif "-" in right[1:]:
            split_at = right[1:].find("-") + 1
            frac, tz_suffix = right[:split_at], right[split_at:]
        frac = frac[:6].ljust(6, "0")
        normalized = f"{left}.{frac}{tz_suffix}"

    try:
        return datetime.fromisoformat(normalized), None
    except ValueError:
        warning = (
            f"{case_id}:timing_log_unparseable_datetime:"
            f"{source_path.name}:row={row_number}:field={field_name}:value={text}"
        )
        return None, warning


def parse_timing_log_csv(path: Path, case_id: str) -> tuple[list[TimingLogEntry], list[str]]:
    if not path.exists():
        return [], []

    warnings: list[str] = []

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
    except OSError as exc:
        raise TimingLogParseError(case_id, path, f"Failed to read timing-log CSV: {exc}") from exc

    if not rows:
        raise TimingLogParseError(case_id, path, "Timing-log CSV is empty.")

    header = [col.strip() for col in rows[0]]
    has_explicit = all(name in header for name in _EXPLICIT_COLUMNS)
    has_positional = len(header) >= 5
    if not has_explicit and not has_positional:
        raise TimingLogParseError(
            case_id,
            path,
            "Timing-log CSV missing expected columns. "
            "Expected explicit headers (Events, TimeSTART, TimeEND) "
            "or at least 5 columns for positional fallback.",
        )

    if len(rows) == 1:
        raise TimingLogParseError(case_id, path, "Timing-log CSV has header only and no data rows.")

    entries: list[TimingLogEntry] = []
    if has_explicit:
        events_idx = header.index("Events")
        start_idx = header.index("TimeSTART")
        end_idx = header.index("TimeEND")
    else:
        events_idx = _POS_EVENT_IDX
        start_idx = _POS_START_IDX
        end_idx = _POS_END_IDX

    for line_idx, row in enumerate(rows[1:], start=2):
        if len(row) <= max(events_idx, start_idx, end_idx):
            raise TimingLogParseError(
                case_id,
                path,
                f"Row {line_idx} does not contain required timing-log columns.",
            )

        label_raw = row[events_idx]
        start_raw = row[start_idx]
        end_raw = row[end_idx]

        if label_raw.strip() == "" and start_raw.strip() == "" and end_raw.strip() == "":
            warnings.append(f"{case_id}:timing_log_blank_row:{path.name}:row={line_idx}")
            continue

        time_start, start_warning = _parse_optional_datetime(
            start_raw,
            case_id=case_id,
            source_path=path,
            row_number=line_idx,
            field_name="TimeSTART",
        )
        if start_warning:
            warnings.append(start_warning)

        time_end, end_warning = _parse_optional_datetime(
            end_raw,
            case_id=case_id,
            source_path=path,
            row_number=line_idx,
            field_name="TimeEND",
        )
        if end_warning:
            warnings.append(end_warning)

        entries.append(
            TimingLogEntry(
                case_id=case_id,
                source_file=path.resolve(),
                row_number=line_idx,
                label_text=label_raw,
                time_start_raw=start_raw if start_raw.strip() != "" else None,
                time_end_raw=end_raw if end_raw.strip() != "" else None,
                time_start=time_start,
                time_end=time_end,
            )
        )

    return entries, warnings


def _xlsx_header(ws: Any, *, case_id: str, path: Path) -> tuple[int, int, int, int] | None:
    """Return the unique EVENT/START/END header coordinates for one sheet."""
    matches: list[tuple[int, int, int, int]] = []
    max_row = min(int(ws.max_row or 0), _XLSX_SCAN_ROWS)
    max_column = min(int(ws.max_column or 0), _XLSX_SCAN_COLUMNS)
    for row_number in range(1, max_row + 1):
        by_name: dict[str, int] = {}
        for column_number in range(1, max_column + 1):
            value = ws.cell(row_number, column_number).value
            normalized = str(value).strip().upper() if value is not None else ""
            if normalized in _XLSX_HEADER_NAMES:
                if normalized in by_name:
                    raise TimingLogParseError(
                        case_id,
                        path,
                        f"Worksheet {ws.title!r} row {row_number} contains duplicate {normalized!r} headers.",
                    )
                by_name[normalized] = column_number
        if all(name in by_name for name in _XLSX_HEADER_NAMES):
            matches.append(
                (
                    row_number,
                    by_name["EVENT"],
                    by_name["START"],
                    by_name["END"],
                )
            )
    if len(matches) > 1:
        raise TimingLogParseError(
            case_id,
            path,
            f"Worksheet {ws.title!r} contains multiple EVENT/START/END header rows.",
        )
    return matches[0] if matches else None


def _clock_from_xlsx_value(value: Any, *, epoch: datetime) -> time | None:
    """Return a clock value for Excel time-only representations."""
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool) and 0 <= float(value) < 1:
        converted = from_excel(float(value), epoch=epoch)
        if isinstance(converted, time):
            return converted
        if isinstance(converted, datetime):
            return converted.time()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for format_text in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                return datetime.strptime(text, format_text).time()
            except ValueError:
                continue
    return None


def _xlsx_raw_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="microseconds")
    if isinstance(value, time):
        return value.isoformat(timespec="microseconds")
    text = str(value).strip()
    return text or None


def _parse_xlsx_timestamp(
    value: Any,
    *,
    case_id: str,
    source_path: Path,
    row_number: int,
    field_name: str,
    reference_date: date | None,
    current_date: date | None,
    previous_timestamp: datetime | None,
    epoch: datetime,
) -> tuple[datetime | None, date | None, str | None]:
    """Parse one XLSX datetime/clock cell and apply an explicit date anchor."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, current_date, None
    if isinstance(value, datetime):
        return value, value.date(), None
    if isinstance(value, date):
        parsed = datetime.combine(value, time.min)
        return parsed, value, None

    clock_value = _clock_from_xlsx_value(value, epoch=epoch)
    if clock_value is not None:
        anchor_date = current_date or reference_date
        if anchor_date is None:
            raise TimingLogParseError(
                case_id,
                source_path,
                f"Row {row_number} field {field_name} contains a time-only value but no treatment-date anchor is available.",
            )
        parsed = datetime.combine(anchor_date, clock_value)
        warning = None
        if previous_timestamp is not None and parsed < previous_timestamp - timedelta(hours=12):
            anchor_date += timedelta(days=1)
            parsed = datetime.combine(anchor_date, clock_value)
            warning = (
                f"{case_id}:timing_log_midnight_rollover:{source_path.name}:"
                f"row={row_number}:field={field_name}:date={anchor_date.isoformat()}"
            )
        return parsed, anchor_date, warning

    if isinstance(value, str):
        parsed, warning = _parse_optional_datetime(
            value,
            case_id=case_id,
            source_path=source_path,
            row_number=row_number,
            field_name=field_name,
        )
        if parsed is not None:
            return parsed, parsed.date(), None
        return None, current_date, warning

    warning = (
        f"{case_id}:timing_log_unparseable_datetime:{source_path.name}:"
        f"row={row_number}:field={field_name}:value={value}"
    )
    return None, current_date, warning


def parse_timing_log_xlsx(
    path: Path,
    case_id: str,
    *,
    reference_datetime: datetime | None,
) -> tuple[list[TimingLogEntry], list[str]]:
    """
    Parse a read-only XLSX timing log into the existing enrichment contract.

    Excel clock-only cells are anchored to the date of the first normalized
    case event. A clock regression greater than 12 hours is treated as a
    midnight rollover and reported. Source workbooks are never modified.
    """
    if not path.exists():
        return [], []
    try:
        with runtime_warnings.catch_warnings():
            runtime_warnings.filterwarnings(
                "ignore",
                message="Data Validation extension is not supported and will be removed",
                category=UserWarning,
                module=r"openpyxl\.worksheet\._reader",
            )
            workbook = load_workbook(
                path,
                read_only=False,
                data_only=True,
                keep_links=False,
            )
    except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException) as exc:
        raise TimingLogParseError(case_id, path, f"Failed to read timing-log XLSX: {exc}") from exc

    try:
        sheet_matches: list[tuple[Any, tuple[int, int, int, int]]] = []
        for worksheet in workbook.worksheets:
            header = _xlsx_header(worksheet, case_id=case_id, path=path)
            if header is not None:
                sheet_matches.append((worksheet, header))
        if len(sheet_matches) != 1:
            raise TimingLogParseError(
                case_id,
                path,
                "Expected exactly one worksheet with an EVENT/START/END header; "
                f"found {len(sheet_matches)}.",
            )

        worksheet, (header_row, event_column, start_column, end_column) = sheet_matches[0]
        populated_rows = [
            row_number
            for row_number in range(header_row + 1, int(worksheet.max_row or 0) + 1)
            if any(
                worksheet.cell(row_number, column_number).value is not None
                for column_number in (event_column, start_column, end_column)
            )
        ]
        if not populated_rows:
            raise TimingLogParseError(case_id, path, "Timing-log XLSX has a header but no data rows.")

        warnings: list[str] = []
        entries: list[TimingLogEntry] = []
        current_date = reference_datetime.date() if reference_datetime is not None else None
        previous_timestamp: datetime | None = None
        for row_number in range(header_row + 1, max(populated_rows) + 1):
            label_value = worksheet.cell(row_number, event_column).value
            start_value = worksheet.cell(row_number, start_column).value
            end_value = worksheet.cell(row_number, end_column).value
            if label_value is None and start_value is None and end_value is None:
                warnings.append(f"{case_id}:timing_log_blank_row:{path.name}:row={row_number}")
                continue

            label_text = "" if label_value is None else str(label_value)
            time_start, current_date, start_warning = _parse_xlsx_timestamp(
                start_value,
                case_id=case_id,
                source_path=path,
                row_number=row_number,
                field_name="START",
                reference_date=reference_datetime.date() if reference_datetime is not None else None,
                current_date=current_date,
                previous_timestamp=previous_timestamp,
                epoch=workbook.epoch,
            )
            if start_warning:
                warnings.append(start_warning)
            if time_start is not None:
                previous_timestamp = time_start

            time_end, current_date, end_warning = _parse_xlsx_timestamp(
                end_value,
                case_id=case_id,
                source_path=path,
                row_number=row_number,
                field_name="END",
                reference_date=reference_datetime.date() if reference_datetime is not None else None,
                current_date=current_date,
                previous_timestamp=previous_timestamp,
                epoch=workbook.epoch,
            )
            if end_warning:
                warnings.append(end_warning)
            if time_end is not None:
                previous_timestamp = time_end

            entries.append(
                TimingLogEntry(
                    case_id=case_id,
                    source_file=path.resolve(),
                    row_number=row_number,
                    label_text=label_text,
                    time_start_raw=_xlsx_raw_text(start_value),
                    time_end_raw=_xlsx_raw_text(end_value),
                    time_start=time_start,
                    time_end=time_end,
                )
            )
        return entries, warnings
    finally:
        workbook.close()


def parse_timing_log(
    path: Path,
    case_id: str,
    *,
    reference_datetime: datetime | None = None,
) -> tuple[list[TimingLogEntry], list[str]]:
    """Dispatch a supported timing-log file to its deterministic parser."""
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return parse_timing_log_csv(path, case_id)
    if suffix == ".xlsx":
        return parse_timing_log_xlsx(
            path,
            case_id,
            reference_datetime=reference_datetime,
        )
    raise TimingLogParseError(
        case_id,
        path,
        f"Unsupported timing-log file extension {path.suffix!r}; expected .csv or .xlsx.",
    )


def timing_log_source_type(path: Path | None) -> str:
    """Return stable analytical-input provenance for a timing-log path."""
    if path is not None and path.suffix.casefold() == ".xlsx":
        return "timing_log_xlsx"
    return "timing_log_csv"
